import os
import itertools
from datetime import date
import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.styles.borders import Border, Side


class bcolors:
    OKGREEN = '\033[92m'
    FAIL = '\033[91m'
    ENDC = '\033[0m'


def section_breaker(custum_list:list, start:int, end:int):
    return(custum_list[start:end])


def order_fix(pd_order, smart_dict):
    """rearranges the order of smart_dict based on the order of pd_order"""
    tmp_dict = {}
    for item in pd_order:
        for disk, smart in smart_dict.items():
            for line in smart:
                if '-d megaraid,' in line:
                    dev_id = line[43:].strip()
                    if item == dev_id:
                        tmp_dict[item] = smart_dict[disk]

    return(tmp_dict)


def num_of_lines(file_obj):
    with open (file_obj, 'r') as f_obj:
        return(len(f_obj.readlines()))


def excel_maker(disk_type, pd_params, smart_params_r, smart_params_v):

    E_S = f"{pd_params['EncID']}/{pd_params['slotNum']}"
    # print(E_S)

    if disk_type == 'hdd':
        rows = (
            (E_S, smart_params_r['Serial'],'Smart Health Status', smart_params_r['SMART Health Status']),
            (None, None,'Elements in grown defect list', smart_params_r['Elements in grown defect list']),
            (None, None,'Total Uncorrected Errors (Read + Write)', smart_params_r['Uncorrected_Errs']),
            (None, None,'Accumulated start-stop cycles', smart_params_r['Accumulated start-stop cycles']),
            (None, None,'Accumulated load-unload cycles', smart_params_r['Accumulated load-unload cycles'])
        )
        height = 5

    elif disk_type == 'micron':
        rows = (
            (E_S, smart_params_r['Serial'],'Raw Read Error Rate', smart_params_v['Raw_Read_Error_Rate'], smart_params_r['Raw_Read_Error_Rate']),
            (None, None, 'Reallocated_Sector_Ct', smart_params_v['Reallocated_Sector_Ct'], smart_params_r['Reallocated_Sector_Ct']),
            (None, None, 'Reported Uncorrect', smart_params_v['Reported_Uncorrect'], smart_params_r['Reported_Uncorrect']),
            (None, None, 'Hardware_ECC_Recovered', smart_params_v['Hardware_ECC_Recovered'], smart_params_r['Hardware_ECC_Recovered']),
            (None, None, 'Unused_Rsvd_Blk_Cnt_Tot', smart_params_v['Unused_Rsvd_Blk_Cnt_Tot'], smart_params_r['Unused_Rsvd_Blk_Cnt_Tot']),
            (None, None, 'Total_LBAs_Written', smart_params_v['246 Unknown_Attribute'], smart_params_r['246 Unknown_Attribute'])
            )
        height = 6

    elif disk_type == 'samsung':
        rows = (
            (E_S, smart_params_r['Serial'], 'Reallocated_Sector_Ct', smart_params_v['Reallocated_Sector_Ct'], smart_params_r['Reallocated_Sector_Ct']),
            (None, None, 'Power On', smart_params_v['Power_On'], smart_params_r['Power_On']),
            (None, None, 'Wear_Leveling_Count', smart_params_v['Wear_Leveling_Count'], smart_params_r['Wear_Leveling_Count']),
            (None, None, 'Used_Rsvd_Blk_Cont_Tot', smart_params_v['Used_Rsvd_Blk_Cnt_Tot'], smart_params_r['Used_Rsvd_Blk_Cnt_Tot']),
            (None, None, 'Runtime Bad Block', smart_params_v['Runtime_Bad_Block'], smart_params_r['Runtime_Bad_Block']),
            (None, None, 'Reported Uncorrect', smart_params_v['Reported_Uncorrect'], smart_params_r['Reported_Uncorrect']),
            (None, None, 'Hardware ECC Recovered', smart_params_v['Hardware_ECC_Recovered'], smart_params_r['Hardware_ECC_Recovered']),
            (None, None, 'Total LBAs Written', smart_params_v['Total_LBAs_Written'], smart_params_r['Total_LBAs_Written'])
        )
        height = 8


    file_list = [file.name for file in os.scandir()]
    file_exist = any('Smart_Parameter' in file for file in file_list)
    if not file_exist:
        workbook = openpyxl.Workbook()
        sheet = workbook.create_sheet("Smart_Parameters")
        workbook.save(f"Smart_Parameters-{date.today().strftime('%B-%d-%Y')}.xlsx")

    for file in os.scandir():
        if 'Smart_Parameters' in file.name:
            final_file = file.name
            final_file = os.path.join(os.getcwd(), final_file)

    workbook = openpyxl.load_workbook(final_file)
    sheet = workbook["Smart_Parameters"]

    for row in rows:
        sheet.append(row)
    sheet.append((None, None, None, None, None, ' '))
    # sheet.merge_cells(f'{}:{}')

    workbook.save(f"Smart_Parameters-{date.today().strftime('%B-%d-%Y')}.xlsx")
    workbook = openpyxl.load_workbook(final_file)
    if 'Sheet' in workbook.sheetnames:
        workbook.remove(workbook['Sheet'])
        workbook.save(final_file)


def get_cells(start, end):
    print(start)
    print(end)


def create_border(cell_range, sheet):
    """Makes thin borders around cells"""
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for row in sheet[cell_range]:
        for cell in row:
            cell.border = thin_border


def cell_borders(disk_types):
    for file in os.scandir():
        if 'Smart_Parameters' in file.name:
            excel_file = file.name
            excel_file = os.path.join(os.getcwd(), excel_file)

    """Makes thin borders around cells"""
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Smart_Parameters"]

    row_index = 0
    # print(disk_types)
    for type in disk_types:
        row_index += 1
        match(type):
            case 'hdd':
                create_border(f"A{row_index}:D{row_index+4}", sheet)
                row_index += 5
            case 'samsung':
                create_border(f"A{row_index}:E{row_index+7}", sheet)
                row_index += 8
            case 'micron':
                create_border(f"A{row_index}:E{row_index+5}", sheet)
                row_index += 6
        workbook.save(excel_file)


def excel_modifier():
    """Sets the column width sizes of each sheet to the appropriate value"""
    # Gets the excel file
    for file in os.scandir():
        if 'Smart_Parameters' in file.name:
            excel_file = file.name
            excel_file = os.path.join(os.getcwd(), excel_file)
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook["Smart_Parameters"]
    center_align = Alignment(horizontal='center', vertical='center')


    dims = {}
    for row in sheet.rows:
        for cell in row:
            cell.alignment = center_align
            if cell.value:
                dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

    for col, value in dims.items():
        sheet.column_dimensions[col].width = value + 4


    a_value_indexes = []
    b_value_indexes = []
    c_empty_indexes = []
    for index, column in enumerate(sheet.columns):
        if 'A' in get_column_letter(index + 1):
            a_value_indexes.extend((cell.row, cell.column) for cell in column if cell.value)
        if 'B' in get_column_letter(index + 1):
            b_value_indexes.extend((cell.row, cell.column) for cell in column if cell.value)
        if 'C' in get_column_letter(index + 1):
            c_empty_indexes.extend(cell.row - 1 for cell in column if not cell.value)

        # Formatting Columns D & E as number
        if 'D' or 'E' in get_column_letter(index + 1):
            for cell in column:
                cell.number_format = '0'

    for indice, cell in zip(c_empty_indexes, a_value_indexes):
        sheet.merge_cells(start_row=cell[0], end_row=indice, start_column=cell[1], end_column=cell[1])
    for indice, cell in zip(c_empty_indexes, b_value_indexes):
        sheet.merge_cells(start_row=cell[0], end_row=indice, start_column=cell[1], end_column=cell[1])

    workbook.save(excel_file)


class Disk():
    def __init__(self, pd_slice, smart_slice):
        self.disk_type = ''
        self.disk_types_list = []
        self.pd_parameters = {
            'devID' : '',
            'EncID' : '',
            'slotNum' : '',
            'firmState' : '',
            'inquiry' : '',
            'diskType' : '',
        }
        self.blank = False
        self.smartparams_raw = {}
        self.smartparams_value = {}

        self.pd_slice = pd_slice
        self.smart_slice = smart_slice

        self.check_blank()
        self.get_pd_params()
        self.disk_type_separation()

        # print(self.blank)
        if not self.blank:
            self.get_smart_params()

        self.disk_info()


    def check_blank(self):
        for line in self.smart_slice:
            if 'Device does not support Self Test logging' in line:
                # print(line)
                self.blank = True

    def get_pd_params(self):
        for line in self.pd_slice:
            if 'Device Id' in line:
                self.pd_parameters['devID'] = int(line[10:].strip())
            if 'Enclosure Device ID' in line:
                self.pd_parameters['EncID'] = int(line[20:].strip())
            if 'Slot Number' in line:
                self.pd_parameters['slotNum'] = int(line[12:].strip())
            if 'firmware state' in line.lower():
                self.pd_parameters['firmState'] = line[15:].strip()
            if 'Inquiry Data' in line:
                self.pd_parameters['inquiry'] = line[13:].strip()
            if 'Media Type' in line:
                self.pd_parameters['diskType'] = line[12:].strip()
        # Disk type

    def get_smart_params(self):
        def samsung():
            search_terms = ['Reallocated_Sector_Ct', 'Power_On',
             'Wear_Leveling_Count', 'Used_Rsvd_Blk_Cnt_Tot',
             'Runtime_Bad_Block', 'Reported_Uncorrect',
             'Hardware_ECC_Recovered','Total_LBAs_Written', 'Serial']
            for line, term in itertools.product(self.smart_slice, search_terms):
                if term in line:
                    if term != 'Serial':
                        self.smartparams_value[term] = int(line.split()[3])
                        self.smartparams_raw[term] = int(line.split()[9])
                    else:
                        self.smartparams_raw[term] = line.split()[2]


        def micron():
            search_terms = ['Raw_Read_Error_Rate', 'Reallocated_Sector_Ct',
             'Reported_Uncorrect', 'Hardware_ECC_Recovered',
             'Unused_Rsvd_Blk_Cnt_Tot', '246 Unknown_Attribute', 'Serial']
            for line, term in itertools.product(self.smart_slice, search_terms):
                if term in line:
                    if term != 'Serial':
                        self.smartparams_value[term] = int(line.split()[3])
                        self.smartparams_raw[term] = int(line.split()[9])
                    else:
                        self.smartparams_raw[term] = line.split()[2]


        def hdd():
            search_terms = ['SMART Health Status', 'Elements in grown defect list',
             'Accumulated start-stop cycles', 'Accumulated load-unload cycles', 'Serial']
            total_uncorrected_errs = 0
            for line, term in itertools.product(self.smart_slice, search_terms):
                if term in line:
                    if term == 'Serial':
                        self.smartparams_raw[term] = line.split()[2]
                    elif term == 'SMART Health Status':
                        self.smartparams_raw[term] = line.split()[3]
                    elif term == 'Elements in grown defect list':
                        self.smartparams_raw[term] = int(line.split()[5])
                    else:
                        self.smartparams_raw[term] = int(line.split()[3])
                if 'read:' in line:
                    total_uncorrected_errs += int(line.split()[7])
                if 'write:' in line:
                    total_uncorrected_errs += int(line.split()[7])
            self.smartparams_raw['Uncorrected_Errs'] = total_uncorrected_errs


        if self.disk_type == 'samsung':
            samsung()
        elif self.disk_type == 'micron':
            micron()
        elif self.disk_type == 'hdd':
            hdd()

    def disk_type_separation(self):
        if 'samsung' in self.pd_parameters['inquiry'].lower():
            self.disk_type = 'samsung'
        if 'micron' in self.pd_parameters['inquiry'].lower():
            self.disk_type = 'micron'
        if 'hard' in self.pd_parameters['diskType'].lower():
            self.disk_type = 'hdd'

    def disk_info(self):
        if not self.blank:
            print(f"{bcolors.OKGREEN}Disk {self.smartparams_raw['Serial']} is {self.disk_type.upper()}", end=' ')
            print(f"and is located at {self.pd_parameters['EncID']}:{self.pd_parameters['slotNum']}{bcolors.ENDC}")
        else:
            print(f"{bcolors.FAIL}Disk {self.pd_parameters['inquiry']}", end=' ')
            print(f"located at {self.pd_parameters['EncID']}:{self.pd_parameters['slotNum']} is", end=' ')
            print(f"{self.pd_parameters['firmState']}.")
            print(f"Smart data isn't available for this disk.{bcolors.ENDC}")


class Chaos():
    def __init__(self, pdlist, smart):

        self.pd_section_indices = []
        # List of pdlist lines
        self.pdlist_list = []
        self.smart_list = []
        # {disk-no: [list of pdlist lines for the Disk-#num]}
        self.pdlist_dict = {}
        self.smart_dict = {}
        self.num_of_disks = len(self.pd_section_indices)
        self.pdlist = pdlist
        self.smart = smart
        self.smart_last_line = num_of_lines(self.smart)
        self.pdlist_last_line = num_of_lines(self.pdlist)

        # Read pdlist and smart files and extend pdlist and smart lists
        with open(self.pdlist, 'r') as f_obj:
            self.pdlist_list.extend(iter(f_obj))

        with open (self.smart, 'r') as f_obj:
            self.smart_list.extend(iter(f_obj))

    def pdlist_slicer(self):
        """Get pdlist sections line indices"""
        self.pd_section_indices = [i for i, x in enumerate(self.pdlist_list) if "Enclosure Device ID" in x]

        for i in range(len(self.pd_section_indices)):
            if i + 1 < len(self.pd_section_indices):
                self.pdlist_dict[f'Disk-{i}'] = section_breaker(self.pdlist_list, self.pd_section_indices[i], self.pd_section_indices[i+1])
            else:
                self.pdlist_dict[f'Disk-{i}'] = section_breaker(self.pdlist_list, self.pd_section_indices[i], self.pdlist_last_line)

    def smart_slicer(self):
        """Get Smart sections line indices"""
        self.smart_section_indices = [i for i, x in enumerate(self.smart_list) if "Command smartctl" in x]

        for i in range(len(self.smart_section_indices)):
            if i + 1 != len(self.smart_section_indices):
                self.smart_dict[f'Disk-{i}'] = section_breaker(self.smart_list, self.smart_section_indices[i], self.smart_section_indices[i+1])
            else:
                self.smart_dict[f'Disk-{i}'] = section_breaker(self.smart_list, self.smart_section_indices[i], self.smart_last_line)

    def compare_order(self):
        pd_id_order = []
        for pd in self.pdlist_dict.values():
            pd_id_order.extend(item[11:].strip() for item in pd if 'Device Id:' in item)

        smart_id_order = []
        for smart in self.smart_dict.values():
            smart_id_order.extend(item[43:].strip() for item in smart if '-d megaraid,' in item)

        self.smart_dict = order_fix(pd_id_order, self.smart_dict)


for file in os.scandir(os.getcwd()):
    if 'pdlist' in file.name.lower():
        pdlist = file.name
    if 'smarts' in file.name.lower():
        smart = file.name

smart = Chaos(pdlist, smart)
smart.pdlist_slicer()
smart.smart_slicer()
smart.compare_order()


disks = {disk: Disk(pd_data, smart_data) for (disk, pd_data),
 (disk_2, smart_data) in zip(smart.pdlist_dict.items(), smart.smart_dict.items())}

disk_types_list = []

for disk in disks.values():
    if not disk.blank:
        excel_maker(disk.disk_type, disk.pd_parameters
        , disk.smartparams_raw, disk.smartparams_value)
        disk_types_list.append(disk.disk_type)

print(disk_types_list)

cell_borders(disk_types_list)
excel_modifier()
